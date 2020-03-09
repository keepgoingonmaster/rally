import openpyxl
from openpyxl.utils import get_column_letter
import sys
import re


def retrieveTestcaseResultFromSheet(excelname):
    wb = openpyxl.load_workbook(excelname)
    sheet = wb['Sheet']

    fields = []

    for column in range(1, sheet.max_column + 1):
        columnLetter = get_column_letter(column)
        fields.append(sheet[columnLetter + str(1)].value)

    testcaseResult = []

    for row in range(2, sheet.max_row + 1):
        eachTestCaseResult = {}

        for column, field in enumerate(fields):
            columnLetter = get_column_letter(column + 1)
            cellValue = sheet[columnLetter + str(row)].value
            if column == 0:
                eachTestCaseResult.setdefault(cellValue, {})
                key = cellValue
                continue
            eachTestCaseResult[key][field] = cellValue

        testcaseResult.append(eachTestCaseResult)

    return testcaseResult

def parseConnectionConfiguration(configureFile):

    info = {}
    with open(configureFile, 'r') as fileHandle:
        line = fileHandle.readline()
        if line.match("apikey"):
            info['apikey'] = line.match('').group(0)
        elif line.match('server'):
            info['server'] = line.match('server').group(0)
        elif line.match('server'):
            info['username'] = line.match('server').group(0)
        elif line.match('server'):
            info['password'] = line.match('server').group(0)
        elif line.match('server'):
            info['workspace'] = line.match('server').group(0)
        elif line.match('server'):
            info['project'] = line.match('server').group(0)


# if __name__ == '__main__':
#     retrieveTestcaseResultFromSheet();