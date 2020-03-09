# !/usr/bin/env python

#################################################################################################
#
# getitem.py -- Get info for a specific instance of a Rally type
#               identified either by an OID or a FormattedID value
#
USAGE = """
Usage: getitem.py <entity_name> <OID | FormattedID>    
"""
#################################################################################################

import sys
import re
import string
from datetime import datetime
import coloredlogs, logging, os
import json

from pyral import Rally, rallyWorkset

# import RallyUtil
from example import RallyUtil
import openpyxl
from parseExcelEnhancement import retrieveTestcaseResultFromSheet
import xlsxwriter
import openpyxl
import xlrd
from collections import deque

logger = logging.getLogger('mylogger')
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

fh = logging.FileHandler(os.path.join(os.path.dirname(os.path.abspath("__file__")), 'testerror.log'))  # file log
ch = logging.StreamHandler()

fh.setFormatter(formatter)
ch.setFormatter(formatter)

logger.addHandler(fh)
logger.addHandler(ch)

#################################################################################################
#################################################################################################
def createlink(excelname, collectionitems, urlprefix, sheetname1='Hyperlinks', sheetname2 ='result'):
    # sheetname1 = 'Hyperlinks'
    # sheetname2 = 'result'
    workbook = xlsxwriter.Workbook(excelname)
    worksheet = workbook.add_worksheet(sheetname1)
    titlelist = ["TestCaseID", "Description"]

    # Format the first column
    worksheet.set_column('A:A', 11)

    # Add a sample alternative link format.
    head_format = workbook.add_format({
        # 'font_color' : 'red',
        'bold' : 1,
        # 'underline' : 1,
        'align': "center",
        'border': 1,
        'font_size' : 12,
    })
    # head_format.set_border()
    # head_format.set_align('center')

    cell_format = workbook.add_format({
        'border': 1
    })
    # cell_format.set_border()

    idx = 1
    for eachitem in collectionitems :
        worksheet.write_url('A'+str(idx), urlprefix+str(eachitem.oid), string=eachitem.FormattedID)
        worksheet.write('B' + str(idx), eachitem.Name)
        idx = idx + 1

    resultsheet = workbook.add_worksheet(sheetname2)
    # workbook.close()

    wbread = openpyxl.load_workbook(excelname)
    sheet = wbread[sheetname1]
    resultsheet.write(0, 0, titlelist[0], head_format)
    resultsheet.write(0, 1, titlelist[1], head_format)
    for row in range(1, sheet.max_row+1) :
        for col in range(1, sheet.max_column+1) :
            if col == 1:
                hyperlink = sheet.cell(row, col).hyperlink.target + "#" + sheet.cell(row, col).hyperlink.location
                if not hyperlink:
                    logging.error("no url")

                resultsheet.write_url(row, col-1, hyperlink, cell_format, string=sheet.cell(row, col).value)
            else:
                resultsheet.write(row, col-1, sheet.cell(row, col).value, cell_format)

    workbook.close()


def reviewtestcases(rallyUtil, excelname, collectionitems, urlprefix, sheetnamelist, titlelist):
    workbook = xlsxwriter.Workbook(excelname)
    worksheets = []
    for sheetname in sheetnamelist:
        worksheets.append(workbook.add_worksheet(sheetname))


    # Add a sample alternative link format.
    head_format = workbook.add_format({
        # 'font_color' : 'red',
        'bold' : 1,
        # 'underline' : 1,
        'align': "center",
        'border': 1,
        'font_size' : 12,
    })

    cell_format = workbook.add_format({
        'border': 1
    })

    noresult_format = workbook.add_format({
        'font_color': 'red',
        'border' : 1
    })

    idxlist = [2] * (len(worksheets) + 1)

    for worksheet in worksheets:
        worksheet.set_column('A:A', 11)
        for col, title in enumerate(titlelist):
            worksheet.write(0,col,title, head_format)

    # for idxsheet, eachitem in enumerate(collectionitems) :
    for eachitem in collectionitems:
        fillexcelcells(worksheets, idxlist, 0, eachitem, cell_format, noresult_format, False)

        if eachitem.Method == 'Manual' and eachitem.c_TestCaseStatus == 'Active':
            fillexcelcells(worksheets, idxlist, 1, eachitem, cell_format, noresult_format)

        elif eachitem.Method == 'Automated' and eachitem.c_TestCaseStatus == 'Active':
            fillexcelcells(worksheets, idxlist, 2, eachitem, cell_format, noresult_format)

        elif eachitem.Method == 'Manual' and eachitem.c_TestCaseStatus == 'Backlog/Future':
            fillexcelcells(worksheets, idxlist, 3, eachitem, cell_format, noresult_format)

        elif eachitem.Method == 'Automated' and eachitem.c_TestCaseStatus == 'Backlog/Future':
            fillexcelcells(worksheets, idxlist, 4, eachitem, cell_format, noresult_format)

        elif eachitem.Method == 'Manual':
            fillexcelcells(worksheets, idxlist, 5, eachitem, cell_format, noresult_format)

        elif eachitem.Method == 'Automated':
            fillexcelcells(worksheets, idxlist, 6, eachitem, cell_format, noresult_format)

    workbook.close()


def fillexcelcells(worksheets, idxlist, idx, eachitem, cell_format, noresult_format, resultanalysis=True):
    worksheets[idx].write_url('A' + str(idxlist[idx]), urlprefix + str(eachitem.oid), cell_format,
                            string=eachitem.FormattedID)
    worksheets[idx].write('B' + str(idxlist[idx]), eachitem.Name, cell_format)
    worksheets[idx].write('C' + str(idxlist[idx]), eachitem.Type, cell_format)
    worksheets[idx].write('D' + str(idxlist[idx]), eachitem.c_TestCaseStatus, cell_format)
    worksheets[idx].write('E' + str(idxlist[idx]), eachitem.Owner.Name, cell_format)

    if not resultanalysis:
        idxlist[idx] = idxlist[idx] + 1
        return

    testcaseresults = rallyUtil.rally.getCollection(
        "https://rally1.rallydev.com/slm/webservice/v2.0/TestCase/" + str(eachitem.oid) + "/Results")

    dd = deque(testcaseresults, maxlen=1)
    if dd:
        latestresult = dd.pop()
        date = latestresult.Date
        verdict = latestresult.Verdict
        worksheets[idx].write('F' + str(idxlist[idx]), date, cell_format)
        if verdict != 'Pass' :
            worksheets[idx].write('G' + str(idxlist[idx]), verdict, noresult_format)
        else :
            worksheets[idx].write('G' + str(idxlist[idx]), verdict, cell_format)
    else:
        worksheets[idx].write('F' + str(idxlist[idx]), "no result", noresult_format)
        worksheets[idx].write('G' + str(idxlist[idx]), None, cell_format)
        logging.info(f"{eachitem.FormattedID} no result")

    # item = next(testcaseresults, None)
    # if not item :
    #     worksheets[idx].write('F' + str(idxlist[idx]), "no result", noresult_format)
    #     logging.info(f"{eachitem.FormattedID} no result")
    # else :
    #     # while item :
    #     #     date = item.Date
    #     #     verdict = item.Verdict
    #     #     item = next(testcaseresults, None)
    #     worksheets[idx].write('F' + str(idxlist[idx]), date, cell_format)
    #     if verdict != 'Pass':
    #         worksheets[idx].write('G' + str(idxlist[idx]), verdict, noresult_format)
    #     else:
    #         worksheets[idx].write('G' + str(idxlist[idx]), verdict)

    idxlist[idx] = idxlist[idx] + 1

if __name__ == '__main__' :
    logging.basicConfig(level=logging.INFO,
                        # format='%(asctime)s.%(msecs)03d - %(name)s  [%(levelname)s] [%(threadName)s] %(message)s',
                        format='',
                        datefmt='%Y-%m-%d %H:%M:%S',
                        filename='process.log',
                        filemode='w')

    coloredlogs.install()
    logging.addLevelName(logging.WARNING, "\033[1;31m%s\033[1;0m" % logging.getLevelName(logging.WARNING))
    logging.addLevelName(logging.ERROR, "\033[1;41m%s\033[1;0m" % logging.getLevelName(logging.ERROR))
    apikey = "_y0jdQIdhTSmCZ1PuzwmeGW1mYWee5UMNsFwUmFuVsEA"
    server = "rally1.rallydev.com"
    username = "jinewang@microstrategy.com"
    password = ""
    workspace = "MicroStrategy Workspace"
    project = "SR-Intelligence-Kernel"

    rallyUtil = RallyUtil()

    excelname = 'reviewtestcase.xlsx'
    # testcasesList = retrieveTestcaseResultFromSheet(excelname)

    start_time = datetime.now()

    testsetNames = [
        "F26683: Function test",
        "F26683: Performance test",
        "F26683: Scalability test",
        "F26683: Stability test"
    ]
    testTypes = [
        'Functional',
        'Performance',
        'Scalability',
        'Stability'
    ]
    iterationName = "I-7-2020"
    stage = 'Code Freeze'
    owner = 'Jine Wang'
    testsets = []
    for testsetName, testType in zip(testsetNames, testTypes):
        testset = rallyUtil.create_testset(testsetName, iterationName, stage, testType, owner)
        if  testset:
            testsets.append(testset)
        else:
            logging.error("create test set fail")
            for item in testsets:
                rallyUtil.rally.delete(entityName='TestSet', itemIdent=item.FormattedID, project='current',
                                  workspace='current')
            exit(0)

    logging.info(f"create test sets {testsets}")


    testfolderID = "TF8672"
    ident_query = 'FormattedID = "%s"' % testfolderID
    itemResponse = rallyUtil.rally.get('TestFolder', fetch=True, query=ident_query,
                                       workspace=workspace, project=project)

    item = next(itemResponse, None)
    if item:
        testcases = rallyUtil.rally.getCollection('https://rally1.rallydev.com/slm/webservice/v2.0/TestFolder/' + str(item.oid) +'/TestCases')
        for testcase in testcases:
            for idx, testType in enumerate(testTypes):
                if testcase.Type == testType:
                    rallyUtil.add_testcasetotestfolder(testsets[idx], testcase.FormattedID)
            # if testcase.Type == 'Functional':
            #     rallyUtil.add_testcasetotestfolder(testsets[0], testcase.FormattedID)
            # elif testcase.Type == 'Performance':
            #     rallyUtil.add_testcasetotestfolder(testsets[1], testcase.FormattedID)
            # elif testcase.Type == 'Scalability':
            #     rallyUtil.add_testcasetotestfolder(testsets[2], testcase.FormattedID)
            # elif testcase.Type == 'Stability':
            #     rallyUtil.add_testcasetotestfolder(testsets[3], testcase.FormattedID)
           
    end_time = datetime.now()
    # print('Duration: {}'.format(end_time - start_time))
    duration = end_time - start_time
    logging.info(f'Duration: {duration}')


    logging.info("Test is done")
    # print("test is done")



