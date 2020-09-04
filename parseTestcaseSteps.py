import openpyxl
from openpyxl.utils import get_column_letter
import types

# wb = openpyxl.Workbook()
# # wb.save('testcaseResult.xlsx')

# wb = openpyxl.Workbook()
# wb.save('testcaseResultForGivenTestFolder.xlsx')

# wb = openpyxl.Workbook()
# # wb.save('testcase.xlsx')


# def retrieveTestcaseResultFromSheet():
#     wb = openpyxl.load_workbook('testcaseResult.xlsx')
#     sheet = wb['Sheet']
#
#     fields = []
#
#     for column in range(1, sheet.max_column + 1):
#         columnLetter = get_column_letter(column)
#         fields.append(sheet[columnLetter + str(1)].value)
#
#     testcaseResult = []
#
#     for row in range(2, sheet.max_row + 1):
#         eachTestCaseResult = {}
#         for column, field in enumerate(fields):
#             columnLetter = get_column_letter(column+1)
#             eachTestCaseResult[field] = sheet[columnLetter + str(row)].value
#         testcaseResult.append(eachTestCaseResult)
#
#     return testcaseResult

def retrieveTestcaseResultFromSheet(excelname):

    wb = openpyxl.load_workbook(excelname)
    sheet = wb['Sheet']

    scenarios = []
    testcase = {}

    testcaseTitleMarker = 'Test Case'
    prerequisites = "Prerequisites"
    testcaseOwner = "Test Case Creator"
    teststepMarker = "Test Steps"
    inputsInstructions = "inputsInstructions"
    expectedResult = "expectedResult"
    testcasetitle = "testcasetitle"


    for row in range(1, sheet.max_row + 1) :
        columnBValue = sheet['B' + str(row)].value
        if not columnBValue:
            if bool(testcase):
                scenarios.append(testcase)
                testcase.clear()
                teststepFlag = False
            continue

        columnCValue = sheet['C' + str(row)].value
        # try:
        #     # columnBStringToInt = eval(columnBValue)
        #     columnBStringToInt = int(columnBValue)
        # except SyntaxError:
        #     pass

        if type(columnBValue) is types.IntType:
            testcase[inputsInstructions].append(columnCValue)
            testcase[expectedResult].append(columnCValue)
        if columnBValue.find(testcaseTitleMarker) >= 0:
            # testcase.setdefault(testcasetitle, {})
            testcase[testcasetitle] = columnCValue
        elif columnBValue.find(prerequisites) >= 0:
            testcase[prerequisites] = columnCValue
        elif columnBValue.find(testcaseOwner) >= 0:
            testcase[testcaseOwner] = columnCValue
        elif columnBValue.find(teststepMarker) :
            # testcase[testcasetitle].setdefault(inputsInstructions, {})
            testcase[inputsInstructions] = []
            testcase[expectedResult] = []
        # else :
        #     # columnBStringToInt = int(columnBValue)
        #     testcase[inputsInstructions].append(columnCValue)

            # if columnBValue.find(testcaseTitleMarker) >= 0 :
            #     # testcase.setdefault(testcasetitle, {})
            #     testcase[testcasetitle] = columnCValue
            # elif columnBValue.find(prerequisites) >= 0 :
            #     testcase[testcasetitle][prerequisites] = columnCValue
            # elif columnBValue.find(testcaseOwner) >= 0 :
            #     testcase[testcasetitle][testcaseOwner] = columnCValue
            # elif columnBValue.find(teststepMarker) :
            #     # testcase[testcasetitle].setdefault(inputsInstructions, {})
            #     testcase[testcasetitle].setdefault(inputsInstructions, {})
            #     testcase[testcasetitle][inputsInstructions] = []
            # elif columnBStringToInt >= 1 :
            #     testcase[testcasetitle][inputsInstructions].append(columnCValue)
            # testcaseInputs    = []
    # testcaseOutputs   = []
    # inputInstructions = []
    # expectedResult    = []
    # newTestcase = False
    # scenarios   = []
    #
    #
    # for row in range(1, sheet.max_row + 1):
    #     if sheet['B' + str(row)].value == 'Test Steps':
    #         inputInstructions.clear()
    #         expectedResult.clear()
    #         newTestcase = True
    #     elif sheet['B' + str(row)].value == '':
    #         testcaseInputs.append(inputInstructions)
    #         testcaseOutputs.append(expectedResult)
    #         inputInstructions.clear()
    #         expectedResult.clear()
    #         newTestcase = False
    #     elif newTestcase:
    #         inputInstructions.append(sheet['C' + str(row)].value)
    #         expectedResult.append(sheet['D' + str(row)].value)
    #
    # if len(inputInstructions) > 0:
    #     testcaseInputs.append(inputInstructions)
    #     testcaseOutputs.append(expectedResult)
    #
    # scenarios.append(testcaseInputs, testcaseOutputs)
    # fields = []
    #
    # for column in range(1, sheet.max_column + 1):
    #     columnLetter = get_column_letter(column)
    #     fields.append(sheet[columnLetter + str(1)].value)
    #
    # testcaseResult = []
    #
    # for row in range(2, sheet.max_row + 1):
    #     eachTestCaseResult = {}
    #
    #     for column, field in enumerate(fields):
    #         columnLetter = get_column_letter(column + 1)
    #         cellValue = sheet[columnLetter + str(row)].value
    #         if column == 0:
    #             eachTestCaseResult.setdefault(cellValue, {})
    #             key = cellValue
    #             continue
    #         eachTestCaseResult[key][field] = cellValue
    #
    #     testcaseResult.append(eachTestCaseResult)

    return scenarios

if __name__ == '__main__':
    excelname = 'testcaseSteps.xlsx'
    retrieveTestcaseResultFromSheet(excelname);