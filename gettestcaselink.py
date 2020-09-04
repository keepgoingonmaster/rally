# !/usr/bin/env python

#################################################################################################
#
# getitem.py -- Get info for a specific instance of a Rally type
#               identified either by an OID or a FormattedID value
#
USAGE = """
Usage: getitem.py <entity_name> <OID | FormattedID>    
"""
#################################################################################################

import sys
import re
import string
from datetime import datetime
import coloredlogs, logging, os
import json

from pyral import Rally, rallyWorkset

# import RallyUtil
from example import RallyUtil
import openpyxl
from parseExcelEnhancement import retrieveTestcaseResultFromSheet
import xlsxwriter
import openpyxl
import xlrd

logger = logging.getLogger('mylogger')
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')

fh = logging.FileHandler(os.path.join(os.path.dirname(os.path.abspath("__file__")), 'testerror.log'))  # file log
ch = logging.StreamHandler()

fh.setFormatter(formatter)
ch.setFormatter(formatter)

logger.addHandler(fh)
logger.addHandler(ch)

#################################################################################################
#################################################################################################
def autofitexcel(excelname, sheetname='result'):
    import win32com.client as win32
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(excelname)
    ws = wb.Worksheets(sheetname)
    ws.Columns.AutoFit()
    wb.Save()
    excel.Application.Quit()



def createlink(excelname, collectionitems, urlprefix, sheetname1='Hyperlinks', sheetname2 ='result'):
    # sheetname1 = 'Hyperlinks'
    # sheetname2 = 'result'
    workbook = xlsxwriter.Workbook(excelname)
    worksheet = workbook.add_worksheet(sheetname1)
    titlelist = ["TestCaseID", "Description", "Type"]

    # Format the first column
    worksheet.set_column('A:A', 11)

    # Add a sample alternative link format.
    head_format = workbook.add_format({
        # 'font_color' : 'red',
        'bold' : 1,
        # 'underline' : 1,
        'align': "center",
        'border': 1,
        'font_size' : 12,
    })
    # head_format.set_border()
    # head_format.set_align('center')

    cell_format = workbook.add_format({
        'border': 1
    })
    # cell_format.set_border()

    idx = 1
    for eachitem in collectionitems :
        worksheet.write_url('A'+str(idx), urlprefix+str(eachitem.oid), string=eachitem.FormattedID)
        worksheet.write('B' + str(idx), eachitem.Name, cell_format)
        worksheet.write('C' + str(idx), eachitem.Type, cell_format)
        idx = idx + 1

    resultsheet = workbook.add_worksheet(sheetname2)
    # workbook.close()

    wbread = openpyxl.load_workbook(excelname)
    sheet = wbread[sheetname1]
    resultsheet.write(0, 0, titlelist[0], head_format)
    resultsheet.write(0, 1, titlelist[1], head_format)
    for row in range(1, sheet.max_row+1) :
        for col in range(1, sheet.max_column+1) :
            if col == 1:
                hyperlink = sheet.cell(row, col).hyperlink.target + "#" + sheet.cell(row, col).hyperlink.location
                if not hyperlink:
                    logging.error("no url")

                resultsheet.write_url(row, col-1, hyperlink, cell_format, string=sheet.cell(row, col).value)
            else:
                resultsheet.write(row, col-1, sheet.cell(row, col).value, cell_format)


    # Write some hyperlinks
    # worksheet.write_url('A1', 'http://www.python.org/')  # Implicit format.
    # worksheet.write_url('A3', 'http://www.python.org/', string='Python Home')
    # worksheet.write_url('A5', 'http://www.python.org/', tip='Click here')
    # worksheet.write_url('A7', 'http://www.python.org/', red_format)
    # worksheet.write_url('A9', 'mailto:jmcnamara@cpan.org', string='Mail me')

    workbook.close()


if __name__ == '__main__' :
    logging.basicConfig(level=logging.INFO,
                        # format='%(asctime)s.%(msecs)03d - %(name)s  [%(levelname)s] [%(threadName)s] %(message)s',
                        format='',
                        datefmt='%Y-%m-%d %H:%M:%S',
                        filename='process.log',
                        filemode='w')

    coloredlogs.install()
    logging.addLevelName(logging.WARNING, "\033[1;31m%s\033[1;0m" % logging.getLevelName(logging.WARNING))
    logging.addLevelName(logging.ERROR, "\033[1;41m%s\033[1;0m" % logging.getLevelName(logging.ERROR))
    apikey = "_y0jdQIdhTSmCZ1PuzwmeGW1mYWee5UMNsFwUmFuVsEA"
    server = "rally1.rallydev.com"
    username = "jinewang@microstrategy.com"
    password = ""
    workspace = "MicroStrategy Workspace"
    project = "SR-Intelligence-Kernel"

    rallyUtil = RallyUtil()

    excelname = 'testcase.xlsx'
    testcasesList = retrieveTestcaseResultFromSheet(excelname)

    start_time = datetime.now()

    result = {}
    highlevelresult = {}
    for testcaseRow in testcasesList:
       for defectID, testcase_json in testcaseRow.items():
           # eachdefect = dict()
           result.setdefault(defectID, [])
           highlevelresult.setdefault(defectID, [])
           print(defectID)

           ident_query = 'FormattedID = "%s"' % defectID
           ident_query_ = 'FormattedID = "{}"'.format(defectID)
           itemResponse = rallyUtil.rally.get('TestFolder', fetch=True, query=ident_query,
                                                 workspace=workspace)
                                                 # workspace=workspace, project=project)

           item = next(itemResponse, None)
           if not item:
               logging.error(f'{defectID} is not found')
               continue


           collectionitems = rallyUtil.rally.getCollection(
            "https://rally1.rallydev.com/slm/webservice/v2.0/TestFolder/" + str(item.oid) + "/TestCases")

           excelname = "result.xlsx"
           createlink(excelname, collectionitems, "https://rally1.rallydev.com/#/detail/testcase/")
           # autofitexcel(excelname)

           
    end_time = datetime.now()
    print('Duration: {}'.format(end_time - start_time))
    duration = end_time - start_time
    logging.info(f'Duration: {duration}')


    logging.info("Test is done")
    print("test is done")



