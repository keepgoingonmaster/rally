import openpyxl

# wb = openpyxl.Workbook()
# # wb.save('testcaseResult.xlsx')

# wb = openpyxl.Workbook()
# wb.save('testcaseResultForGivenTestFolder.xlsx')

wb = openpyxl.Workbook()
wb.save('testcaseResultForGivenTestset.xlsx')

# wb = openpyxl.Workbook()
# # wb.save('testcaseResult.xlsx')

# wb = openpyxl.Workbook()
# wb.save('testcaseResultForGivenTestFolder.xlsx')

# wb = openpyxl.Workbook()
# # wb.save('testcase.xlsx')


# def retrieveTestcaseResultFromSheet():
#     wb = openpyxl.load_workbook('testcaseResult.xlsx')
#     sheet = wb['Sheet']
#
#     fields = []
#
#     for column in range(1, sheet.max_column + 1):
#         columnLetter = get_column_letter(column)
#         fields.append(sheet[columnLetter + str(1)].value)
#
#     testcaseResult = []
#
#     for row in range(2, sheet.max_row + 1):
#         eachTestCaseResult = {}
#         for column, field in enumerate(fields):
#             columnLetter = get_column_letter(column+1)
#             eachTestCaseResult[field] = sheet[columnLetter + str(row)].value
#         testcaseResult.append(eachTestCaseResult)
#
#     return testcaseResult


def retrieveTestcaseResultFromSheet():
    wb = openpyxl.load_workbook('testcaseResult.xlsx')
    for i in wb.sheetnames:
        print(i)
    sheet = wb['Sheet']

    testcaseResult = {}

    for row in range(2, sheet.max_row + 1):
        testcase          = sheet['A' + str(row)].value
        build             = sheet['B' + str(row)].value
        tester            = sheet['C' + str(row)].value
        verdict           = sheet['D' + str(row)].value
        notes             = sheet['E' + str(row)].value
        productionRelease = sheet['F' + str(row)].value

        testcaseResult.setdefault(testcase, {})
        testcaseResult[testcase]['Build'] = build
        testcaseResult[testcase]['tester'] = tester
        testcaseResult[testcase]['Verdict'] = verdict
        testcaseResult[testcase]['Notes'] = notes
        testcaseResult[testcase]['c_ProductionRelease'] = productionRelease

    return testcaseResult

# if __name__ == '__main__':
#     retrieveTestcaseResultFromSheet();